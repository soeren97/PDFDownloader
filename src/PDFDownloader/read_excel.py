"""Function to read excel file."""

from itertools import islice
from typing import Any, Generator

from openpyxl import load_workbook


def read_excel_rows(
    file_path: str, sheet_name: str
) -> Generator[tuple[Any], None, None]:
    """Read rows of an excel file iteratively.

    Args:
        file_path (str): Location of the excel file.
        sheet_name (str): Name of the sheet.

    Yields:
        Generator[tuple[Any]]: Yielded row.
    """
    # Load the Excel workbook
    wb = load_workbook(filename=file_path, read_only=True)
    # Select the worksheet by name
    ws = wb[sheet_name]

    # Iterate over each row in the worksheet, skipping the first row
    for row in islice(ws.iter_rows(values_only=True), 1, None):
        yield row


def count_excel_rows(file_path: str, sheet_name: str) -> int:
    """Count the number of rows in an Excel file.

    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to count rows from.

    Returns:
        int: Number of rows in the specified sheet.
    """
    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb[sheet_name]
    return sheet.max_row  # type:ignore
